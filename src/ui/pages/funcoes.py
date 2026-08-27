import json
import customtkinter as ctk
from tkinter import messagebox, filedialog
from src.ui.styles import COLORS, get_fonts
from src.utils.paths import get_data_dir

FUNCOES_FILE = get_data_dir() / "funcoes.json"

DEFAULT_FUNCOES = []


def load_funcoes() -> list:
    if FUNCOES_FILE.exists():
        try:
            data = json.loads(FUNCOES_FILE.read_text(encoding="utf-8"))
            return data.get("funcoes", DEFAULT_FUNCOES)
        except Exception:
            pass
    return DEFAULT_FUNCOES


def save_funcoes(funcoes: list):
    FUNCOES_FILE.write_text(
        json.dumps({"funcoes": funcoes}, indent=2, ensure_ascii=False),
        encoding="utf-8"
    )


class FuncoesPage(ctk.CTkFrame):
    ITEMS_PER_PAGE = 20

    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color=COLORS["background"], **kwargs)
        self.funcoes = load_funcoes()
        self.current_page = 1

        self._build_ui()

    def _build_ui(self):
        fonts = get_fonts()

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_propagate(False)

        # Row 0 — Header
        self._header = ctk.CTkFrame(self, fg_color="transparent")
        self._header.grid(row=0, column=0, sticky="ew", padx=20, pady=20)
        self._header.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            self._header, text="Lista de Funcoes",
            font=fonts["title"], text_color=COLORS["primary"]
        ).grid(row=0, column=0, sticky="w")

        ctk.CTkLabel(
            self._header,
            text="Funcoes disponiveis para selecao nos templates de NR",
            font=fonts["small"], text_color=COLORS["muted"]
        ).grid(row=1, column=0, sticky="w", pady=(4, 0))

        btn_frame = ctk.CTkFrame(self._header, fg_color="transparent")
        btn_frame.grid(row=0, column=1, rowspan=2, sticky="e")

        ctk.CTkButton(
            btn_frame, text="Exportar Excel",
            font=fonts["body_bold"], height=32,
            fg_color=COLORS["accent"], hover_color=COLORS["secondary"],
            command=self._export_excel
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            btn_frame, text="Importar Excel",
            font=fonts["body_bold"], height=32,
            fg_color=COLORS["accent"], hover_color=COLORS["secondary"],
            command=self._import_excel
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            btn_frame, text="+ Adicionar",
            font=fonts["body_bold"], height=32,
            fg_color=COLORS["success"], hover_color="#256B28",
            command=self._add_funcao
        ).pack(side="left")

        # Row 1 — Lista (weight=1 preenche resto)
        self.list_frame = ctk.CTkScrollableFrame(self, fg_color=COLORS["surface"], corner_radius=12, height=200)
        self.list_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 5))
        self.list_frame.grid_columnconfigure(0, weight=1)

        # Row 2 — Paginacao custom
        self.pagination_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.pagination_frame.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 5))
        self.pagination_frame.grid_columnconfigure(0, weight=1)

        self.lbl_page = ctk.CTkLabel(
            self.pagination_frame, text="", font=fonts["small"], text_color=COLORS["muted"]
        )
        self.lbl_page.grid(row=0, column=0, sticky="w")

        nav_frame = ctk.CTkFrame(self.pagination_frame, fg_color="transparent")
        nav_frame.grid(row=0, column=1, sticky="e")

        self.btn_prev = ctk.CTkButton(
            nav_frame, text="<", width=30, height=26, font=fonts["small"],
            fg_color=COLORS["primary"], hover_color=COLORS["secondary"],
            command=lambda: self._go_page(self.current_page - 1)
        )
        self.btn_prev.pack(side="left", padx=2)

        self.btn_next = ctk.CTkButton(
            nav_frame, text=">", width=30, height=26, font=fonts["small"],
            fg_color=COLORS["primary"], hover_color=COLORS["secondary"],
            command=lambda: self._go_page(self.current_page + 1)
        )
        self.btn_next.pack(side="left", padx=2)

        self.after(200, lambda: self._fit_scroll_height(0))
        self.after(600, lambda: self._fit_scroll_height(0))

        self._refresh_list()

    def _fit_scroll_height(self, _retry=0):
        self.update_idletasks()
        h = self.winfo_height()
        if h < 200:
            try:
                ph = self.master.winfo_height()
                if ph >= 200:
                    h = ph
            except Exception:
                pass
        if h < 200:
            if _retry < 5:
                self.after(300, lambda r=_retry + 1: self._fit_scroll_height(r))
            return
        header_h = self._header.winfo_reqheight()
        pag_h = self.pagination_frame.winfo_reqheight()
        if min(header_h, pag_h) < 5 and _retry < 5:
            self.after(300, lambda r=_retry + 1: self._fit_scroll_height(r))
            return
        margins = 30
        available = h - header_h - pag_h - margins
        self.list_frame.configure(height=max(available, 150))

    @property
    def total_pages(self) -> int:
        if not self.funcoes:
            return 1
        return max(1, (len(self.funcoes) + self.ITEMS_PER_PAGE - 1) // self.ITEMS_PER_PAGE)

    def _go_page(self, page):
        self.current_page = max(1, min(page, self.total_pages))
        self._refresh_list()

    def _refresh_list(self):
        for widget in self.list_frame.winfo_children():
            widget.destroy()

        fonts = get_fonts()

        if not self.funcoes:
            ctk.CTkLabel(
                self.list_frame, text="Nenhuma funcao cadastrada",
                font=fonts["body"], text_color=COLORS["muted"]
            ).grid(row=0, column=0, pady=40, padx=20)
            self.lbl_page.configure(text="")
            self.btn_prev.configure(state="disabled")
            self.btn_next.configure(state="disabled")
            return

        total = len(self.funcoes)
        start = (self.current_page - 1) * self.ITEMS_PER_PAGE
        end = min(start + self.ITEMS_PER_PAGE, total)
        page_funcoes = self.funcoes[start:end]

        self.lbl_page.configure(text=f"Pagina {self.current_page}/{self.total_pages} | {total} funcoes")
        self.btn_prev.configure(state="normal" if self.current_page > 1 else "disabled")
        self.btn_next.configure(state="normal" if self.current_page < self.total_pages else "disabled")

        for i, func in enumerate(page_funcoes):
            global_idx = start + i
            row = ctk.CTkFrame(self.list_frame, fg_color="transparent")
            row.grid(row=i, column=0, sticky="ew", pady=2, padx=8)
            row.grid_columnconfigure(0, weight=1)

            ctk.CTkLabel(
                row, text=func, font=fonts["body"], text_color=COLORS["text"]
            ).grid(row=0, column=0, sticky="w", padx=12, pady=8)

            btn_frame = ctk.CTkFrame(row, fg_color="transparent")
            btn_frame.grid(row=0, column=1, sticky="e", padx=8, pady=4)

            ctk.CTkButton(
                btn_frame, text="Editar", width=50, height=26,
                font=fonts["small"], fg_color=COLORS["secondary"],
                hover_color=COLORS["primary"],
                command=lambda idx=global_idx: self._edit_funcao(idx)
            ).pack(side="left", padx=2)

            ctk.CTkButton(
                btn_frame, text="Excluir", width=50, height=26,
                font=fonts["small"], fg_color=COLORS["error"],
                hover_color="#B71C1C",
                command=lambda idx=global_idx: self._delete_funcao(idx)
            ).pack(side="left", padx=2)

    def _add_funcao(self):
        self._open_dialog()

    def _edit_funcao(self, index: int):
        self._open_dialog(index)

    def _open_dialog(self, index: int = None):
        fonts = get_fonts()
        is_edit = index is not None

        dialog = ctk.CTkToplevel(self)
        dialog.title("Editar Funcao" if is_edit else "Nova Funcao")
        dialog.geometry("400x200")
        dialog.transient(self)
        dialog.grab_set()
        dialog.resizable(False, False)

        dialog.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width() // 2) - 200
        y = self.winfo_rooty() + (self.winfo_height() // 2) - 75
        dialog.geometry(f"+{x}+{y}")

        form = ctk.CTkFrame(dialog, fg_color="transparent")
        form.pack(fill="both", expand=True, padx=24, pady=24)
        form.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(form, text="Nome da Funcao", font=fonts["body_bold"], text_color=COLORS["text"]).grid(row=0, column=0, sticky="w", pady=(0, 8))

        var = ctk.StringVar(value=self.funcoes[index] if is_edit else "")
        ctk.CTkEntry(form, textvariable=var, font=fonts["body"], height=36, corner_radius=6, placeholder_text="Ex: Tecnico de Seguranca do Trabalho").grid(row=1, column=0, sticky="ew", pady=(0, 16))

        def save():
            nome = var.get().strip()
            if not nome:
                messagebox.showerror("Erro", "Nome e obrigatorio", parent=dialog)
                return
            if is_edit:
                self.funcoes[index] = nome
            else:
                if nome not in self.funcoes:
                    self.funcoes.append(nome)
            save_funcoes(self.funcoes)
            self._refresh_list()
            dialog.destroy()

        btn_frame = ctk.CTkFrame(form, fg_color="transparent")
        btn_frame.grid(row=2, column=0, sticky="ew")

        ctk.CTkButton(btn_frame, text="Cancelar", font=fonts["body"], height=30, fg_color=COLORS["muted"], command=dialog.destroy).pack(side="left")
        ctk.CTkButton(btn_frame, text="Salvar", font=fonts["body_bold"], height=30, fg_color=COLORS["success"], command=save).pack(side="right")

    def _delete_funcao(self, index: int):
        if messagebox.askyesno("Confirmar", f"Excluir '{self.funcoes[index]}'?", parent=self):
            self.funcoes.pop(index)
            save_funcoes(self.funcoes)
            if self.current_page > self.total_pages:
                self.current_page = self.total_pages
            self._refresh_list()

    def _export_excel(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Erro", "openpyxl nao instalado", parent=self)
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="funcoes.xlsx",
            title="Exportar funcoes"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Funcoes"
        ws.append(["Funcao"])
        for f in self.funcoes:
            ws.append([f])
        wb.save(path)
        wb.close()
        messagebox.showinfo("Sucesso", f"Exportado: {len(self.funcoes)} funcoes", parent=self)

    def _import_excel(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Erro", "openpyxl nao instalado", parent=self)
            return

        path = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx")],
            title="Importar funcoes"
        )
        if not path:
            return

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if row and row[0]:
                nome = str(row[0]).strip()
                if nome and nome not in self.funcoes:
                    self.funcoes.append(nome)
                    count += 1
        wb.close()
        save_funcoes(self.funcoes)
        self._refresh_list()
        messagebox.showinfo("Sucesso", f"Importadas {count} novas funcoes", parent=self)
