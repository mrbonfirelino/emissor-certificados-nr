"""Testes ASO + EPI + novos campos de funcionario (v1.11.0).

Rodar: python test_aso_epi.py
"""

import sys
import tempfile
from datetime import date, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

from src.core.employee_repo import EmployeeRepository
from src.core.models import Employee
from src.core.aso_repo import AsoRepository
from src.core.epi_repo import EpiRepository
from src.ui.pages.vencimentos import filter_certs

PASSOS = []


def check(nome, cond):
    PASSOS.append((nome, bool(cond)))
    print(f"[{'OK' if cond else 'FALHOU'}] {nome}")


def make_db(tmp: Path) -> Path:
    tmp.mkdir(parents=True, exist_ok=True)
    db = tmp / "test.db"
    EmployeeRepository(db_path=db)  # cria employees primeiro (JOINs)
    return db


# ── 1. Sequencias ASO/EPI ────────────────────────────────────

def test_sequencias(tmp: Path):
    db = make_db(tmp)
    aso = AsoRepository(db_path=db)
    epi = EpiRepository(db_path=db)
    check("sequencia ASO", aso.next_aso_number() == "ASO-000001"
          and aso.next_aso_number() == "ASO-000002")
    check("sequencia EPI", epi.next_epi_number() == "EPI-000001"
          and epi.next_epi_number() == "EPI-000002")


# ── 2. Expiracao ASO (renovacao substitui) ───────────────────

def test_aso_expiracao(tmp: Path):
    db = make_db(tmp)
    emp_repo = EmployeeRepository(db_path=db)
    aso = AsoRepository(db_path=db)
    emp_repo.create("Joao Pedro", None)
    emp = emp_repo.get_all()[0]

    hoje = date.today().isoformat()
    dois_anos_atras = (date.today() - timedelta(days=730)).isoformat()
    a1 = aso.save("ASO-000001", emp.id, "Periódico", dois_anos_atras, validade_meses=12)
    a2 = aso.save("ASO-000002", emp.id, "Periódico", hoje, validade_meses=12)

    lista = aso.get_asos_with_expiration()
    check("only_latest mantem 1 por funcionario", len(lista) == 1)
    check("vigente e a renovada", lista[0]["cert_number"] == "ASO-000002"
          and lista[0]["nr_code"] == "ASO"
          and lista[0]["descricao_treinamento"] == "Periódico"
          and lista[0]["dias_para_vencer"] > 300)
    todos = aso.get_asos_with_expiration(only_latest=False)
    check("only_latest=False traz todas", len(todos) == 2
          and any(a["status"] == "vencido" for a in todos))

    # outro funcionario nao e afetado
    emp_repo.create("Maria Silva", None)
    maria = emp_repo.get_all()[1]
    aso.save("ASO-000003", maria.id, "Admissional", hoje, validade_meses=6)
    check("dois funcionarios -> 2 vigentes",
          len(aso.get_asos_with_expiration()) == 2)


# ── 3. Novos campos do funcionario ───────────────────────────

def test_campos_funcionario(tmp: Path):
    db = make_db(tmp)
    repo = EmployeeRepository(db_path=db)
    repo.create("Joao Pedro", None, "Eletricista", None, "11999999999",
                data_nascimento="15/03/1990", tipo_sanguineo="a+",
                data_admissao="10/02/2020", registro_ctps="12345/67",
                cnh_ear=True)
    emp = repo.get_all()[0]
    check("create normaliza campos", emp.tipo_sanguineo == "A+"
          and emp.data_admissao == "2020-02-10"
          and emp.registro_ctps == "12345/67" and emp.cnh_ear is True)

    ok = repo.update(emp.id, emp.nome, emp.cpf, emp.funcao,
                     telefone=emp.telefone,
                     tipo_sanguineo=None, limpar_tipo_sanguineo=True,
                     registro_ctps=None, limpar_ctps=True,
                     cnh_ear=False)
    emp2 = repo.get_by_id(emp.id)
    check("update limpa ts/ctps e zera ear", ok and emp2.tipo_sanguineo is None
          and emp2.registro_ctps is None and emp2.cnh_ear is False)
    check("admissao preservada sem flag", emp2.data_admissao == "2020-02-10")

    repo.update(emp.id, emp.nome, emp.cpf, emp.funcao, data_admissao="01/01/2020")
    check("update troca admissao", repo.get_by_id(emp.id).data_admissao == "2020-01-01")


# ── 4. Import/Export colunas F-I ─────────────────────────────

def test_import_export(tmp: Path):
    import openpyxl
    from src.utils.excel_importer import import_employees_from_excel
    from src.utils.excel_exporter import export_employees_to_excel

    db = make_db(tmp)
    repo = EmployeeRepository(db_path=db)

    xlsx_in = tmp / "in.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Nome", "CPF", "Funcao", "Telefone", "Nascimento",
               "Tipo Sangue", "Admissao", "CTPS", "CNH EAR"])
    ws.append(["Ana Souza", "52998224725", "Calheireta", "11988887777",
               "01/02/1990", "O+", "05/01/2021", "99887/66", "Sim"])
    ws.append(["Bruno Lima", "", "", "", "", "", "", "", ""])
    ws.append(["Carla Mota", "", "", "", "", "X+", "", "", ""])
    wb.save(xlsx_in)

    criados, duplicados, erros, detalhes = import_employees_from_excel(str(xlsx_in), repo)
    msgs = " | ".join(str(d) for d in detalhes).lower()
    check("import 2 criados 1 erro", criados == 2 and duplicados == 0
          and erros == 1 and "tipo sanguineo" in msgs)
    ana = repo.search("Ana Souza")[0]
    check("import campos F-I", ana.tipo_sanguineo == "O+"
          and ana.data_admissao == "2021-01-05"
          and ana.registro_ctps == "99887/66" and ana.cnh_ear is True)

    xlsx_out = tmp / "out.xlsx"
    n = export_employees_to_excel(repo, str(xlsx_out))
    wb2 = openpyxl.load_workbook(xlsx_out)
    ws2 = wb2.active
    headers = [c.value for c in ws2[1]]
    check("export headers novos", "Tipo Sanguineo" in headers
          and "Data Admissao" in headers and "Registro CTPS" in headers
          and "CNH EAR" in headers and n == 2)
    ana_row = next(r for r in ws2.iter_rows(min_row=2, values_only=True)
                   if r[1] == "Ana Souza")
    ts_i = headers.index("Tipo Sanguineo")
    ear_i = headers.index("CNH EAR")
    check("export valores", ana_row[ts_i] == "O+" and ana_row[ear_i] == "Sim")


# ── 5. EPI CRUD + status ─────────────────────────────────────

def test_epi_crud(tmp: Path):
    db = make_db(tmp)
    emp_repo = EmployeeRepository(db_path=db)
    epi = EpiRepository(db_path=db)
    emp_repo.create("Joao Pedro", None)
    emp = emp_repo.get_all()[0]

    items = [
        {"ca": "1234", "descricao": "Luva nitrilica", "quantidade": "2",
         "data_entrega": "2026-09-01", "dev_quantidade": "1",
         "dev_data": "2026-09-03"},
        {"ca": "5678", "descricao": "Capacete aba frontal", "quantidade": "1",
         "data_entrega": "2026-09-01", "dev_quantidade": "", "dev_data": ""},
    ]
    fid = epi.save("EPI-000001", emp.id, "2026-09-01", items)
    ficha = epi.get_by_id(fid)
    check("epi save/get", ficha["epi_number"] == "EPI-000001"
          and ficha["status"] == "aberto" and len(ficha["items"]) == 2)
    check("epi items normalizados", ficha["items"][0]["dev_quantidade"] == "1")

    check("get_by_employee", len(epi.get_by_employee(emp.id)) == 1)
    epi.update_items(fid, items[:1])
    check("update_items", len(epi.get_by_id(fid)["items"]) == 1)

    novo = epi.toggle_status(fid)
    check("toggle aberto->fechado", novo == "fechado"
          and epi.get_by_id(fid)["status"] == "fechado")
    check("toggle fechado->aberto", epi.toggle_status(fid) == "aberto")


# ── 6. EPI docs: multiplas versoes ───────────────────────────

def test_epi_docs_versoes(tmp: Path):
    db = make_db(tmp)
    emp_repo = EmployeeRepository(db_path=db)
    epi = EpiRepository(db_path=db)
    emp_repo.create("Maria Silva", None)
    emp = emp_repo.get_all()[0]
    fid = epi.save("EPI-000001", emp.id, "2026-09-01",
                   [{"ca": "1", "descricao": "Bota", "quantidade": "1",
                     "data_entrega": "2026-09-01", "dev_quantidade": "",
                     "dev_data": ""}])

    d1 = epi.add_doc(fid, "EPI-000001_scan1.pdf", b"%PDF-v1", "pdf")
    d2 = epi.add_doc(fid, "EPI-000001_scan2.pdf", b"%PDF-v2", "pdf")
    docs = epi.list_docs(fid)
    check("multiplas versoes convivem", len(docs) == 2 and epi.count_docs(fid) == 2)

    got = epi.get_doc(d1)
    check("get_doc", got and got[0] == fid and got[2] == b"%PDF-v1")

    epi.delete_doc(d2)
    check("delete 1 versao mantem a outra",
          epi.count_docs(fid) == 1 and epi.get_doc(d2) is None)

    try:
        epi.add_doc(fid, "mal.exe", b"MZ", "exe")
        check("exe bloqueado em epi_doc", False)
    except ValueError:
        check("exe bloqueado em epi_doc", True)


# ── 7. Merge vencimentos (formato compativel) ────────────────

def test_merge_vencimentos(tmp: Path):
    db = make_db(tmp)
    emp_repo = EmployeeRepository(db_path=db)
    aso = AsoRepository(db_path=db)
    emp_repo.create("Joao Pedro", None)
    emp = emp_repo.get_all()[0]
    hoje = date.today().isoformat()
    aso.save("ASO-000001", emp.id, "Periódico", hoje, validade_meses=12)

    asos = aso.get_asos_with_expiration()
    fake_certs = [{"nr_code": "NR-35", "funcionario_nome": "Joao Pedro",
                   "funcionario_cpf": "", "dias_para_vencer": 100,
                   "status": "ok", "descricao_treinamento": "NR-35",
                   "data_validade": "2027-01-01", "employee_id": emp.id}]
    combined = fake_certs + asos

    check("chaves exigidas pelo filter_certs presentes",
          all(k in asos[0] for k in
              ("nr_code", "funcionario_nome", "funcionario_cpf", "dias_para_vencer")))
    so_aso = filter_certs(combined, "ASO", "", "all")
    check("filtro NR=ASO isola ASOs", len(so_aso) == 1
          and so_aso[0]["cert_number"] == "ASO-000001")
    check("filtro TODAS traz tudo", len(filter_certs(combined, "TODAS", "", "all")) == 2)


# ── 8. PDFs ASO e EPI ────────────────────────────────────────

def test_pdfs(tmp: Path):
    import fitz
    from src.core.aso_pdf_generator import generate_aso_pdf
    from src.core.epi_pdf_generator import generate_epi_pdf

    tmp.mkdir(parents=True, exist_ok=True)
    emp = Employee(id=1, nome="Joao Pedro Teste", cpf="52998224725",
                   funcao="Eletricista", telefone="11999999999",
                   data_admissao="2020-03-10", tipo_sanguineo="O+")

    p_aso = tmp / "ASO-000001.pdf"
    generate_aso_pdf(str(p_aso), "ASO-000001", emp, "Periódico",
                     date.today().isoformat(), validade_meses=12)
    doc = fitz.open(p_aso)
    txt = "".join(pg.get_text() for pg in doc)
    check("pdf ASO gerado com numero e tipo",
          p_aso.exists() and "ASO-000001" in txt and "Periódico" in txt)

    items = [{"ca": "1234", "descricao": "Luva nitrilica", "quantidade": "2",
              "data_entrega": "01/09/2026", "dev_quantidade": "1",
              "dev_data": "03/09/2026"}]
    p_epi = tmp / "EPI-000001.pdf"
    generate_epi_pdf(str(p_epi), "EPI-000001", emp, "2026-09-01", items)
    doc2 = fitz.open(p_epi)
    txt2 = "".join(pg.get_text() for pg in doc2)
    check("pdf EPI gerado com numero e item",
          p_epi.exists() and "EPI-000001" in txt2 and "Luva nitrilica" in txt2)


def main():
    with tempfile.TemporaryDirectory(prefix="normatech_asoepi_",
                                     ignore_cleanup_errors=True) as td:
        tmp = Path(td)
        test_sequencias(tmp / "t1")
        test_aso_expiracao(tmp / "t2")
        test_campos_funcionario(tmp / "t3")
        test_import_export(tmp / "t4")
        test_epi_crud(tmp / "t5")
        test_epi_docs_versoes(tmp / "t6")
        test_merge_vencimentos(tmp / "t7")
        test_pdfs(tmp / "t8")

    falhas = [n for n, ok in PASSOS if not ok]
    print(f"\n{len(PASSOS) - len(falhas)}/{len(PASSOS)} testes OK")
    if falhas:
        print("FALHARAM:", falhas)
        sys.exit(1)


if __name__ == "__main__":
    main()
