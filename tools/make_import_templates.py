"""Gera os modelos de importacao (.xlsx) na pasta 'MODELOS DE IMPORTACAO'.

Rodar: python tools/make_import_templates.py
"""

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from src.core.aso_repo import ASO_TIPOS

PASTA = ROOT / "MODELOS DE IMPORTACAO"
AZUL = "1B3A5C"
ZEBRA = "EAF0F6"


def _nova_planha(titulo: str, colunas, larguras, linhas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo

    header_fill = PatternFill("solid", fgColor=AZUL)
    header_font = Font(bold=True, color="FFFFFF")
    for col, nome in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col, value=nome)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = larguras[col - 1]
    ws.freeze_panes = "A2"

    zebra_fill = PatternFill("solid", fgColor=ZEBRA)
    for i, linha in enumerate(linhas, start=2):
        for col, valor in enumerate(linha, start=1):
            cell = ws.cell(row=i, column=col, value=valor)
            if i % 2 == 1:
                cell.fill = zebra_fill
    return wb


def _codigos_nr():
    try:
        from src.core.template_loader import load_all_templates
        return sorted(load_all_templates().keys())
    except Exception:
        return ["NR-01", "NR-06", "NR-10", "NR-12", "NR-35"]


def main():
    PASTA.mkdir(exist_ok=True)

    # 1. Funcionarios (A-I)
    wb = _nova_planha(
        "FUNCIONARIOS",
        ["Nome", "CPF", "Funcao", "Telefone", "Data Nascimento",
         "Tipo Sanguineo", "Data Admissao", "Registro CTPS", "CNH EAR"],
        [32, 16, 22, 16, 16, 14, 14, 16, 10],
        [
            ["Joao Pedro da Silva", "529.982.247-25", "Eletricista", "11999998888",
             "15/03/1985", "O+", "10/01/2020", "12345/serie 12", "Sim"],
            ["Maria Souza", "111.444.777-35", "Tecnico de Seguranca", "11988887777",
             "22/07/1990", "A-", "05/03/2021", "", "Nao"],
            ["Carlos Oliveira", "", "Operador de Empilhadeira", "",
             "", "", "", "", ""],
        ],
    )
    wb.save(PASTA / "MODELO FUNCIONARIOS.xlsx")

    # 2. Certificados (A-C)
    wb = _nova_planha(
        "CERTIFICADOS",
        ["Nome", "NR", "Data do Treinamento"],
        [32, 10, 22],
        [
            ["Joao Pedro da Silva", "NR-35", "10/08/2026"],
            ["Maria Souza", "NR-10", "12/08/2026"],
            ["Carlos Oliveira", "NR-12", "15/08/2026"],
        ],
    )
    wb.save(PASTA / "MODELO CERTIFICADOS.xlsx")

    # 3. Cartoes de bloqueio (A-B)
    wb = _nova_planha(
        "CARTOES BLOQUEIO",
        ["Nome", "CPF"],
        [32, 18],
        [
            ["Joao Pedro da Silva", "529.982.247-25"],
            ["Maria Souza", "111.444.777-35"],
        ],
    )
    wb.save(PASTA / "MODELO CARTOES BLOQUEIO.xlsx")

    # 4. ASOs (A-E)
    wb = _nova_planha(
        "ASOS",
        ["Nome", "CPF", "Tipo de ASO", "Data do Exame", "Validade (meses)"],
        [32, 18, 22, 18, 16],
        [
            ["Joao Pedro da Silva", "529.982.247-25", ASO_TIPOS[0], "01/08/2026", 12],
            ["Maria Souza", "111.444.777-35", ASO_TIPOS[1], "05/08/2026", 24],
            ["Carlos Oliveira", "", ASO_TIPOS[0], "", 12],
        ],
    )
    wb.save(PASTA / "MODELO ASO.xlsx")

    # 5. LEIA-ME
    nrs = ", ".join(_codigos_nr())
    leia_me = f"""MODELOS DE IMPORTACAO — NormaTech
=====================================

Como usar: preencha o modelo, apague as linhas de exemplo e importe no app.
A primeira linha (azul) e o cabecalho — nao apague.

1) MODELO FUNCIONARIOS.xlsx  →  aba Funcionarios > botao Importar
   A Nome* | B CPF | C Funcao | D Telefone | E Data Nascimento (dd/mm/aaaa)
   F Tipo Sanguineo (A+, A-, B+, B-, AB+, AB-, O+, O-) | G Data Admissao (dd/mm/aaaa)
   H Registro CTPS | I CNH EAR (Sim/Nao)
   * Nome e obrigatorio. CPF, se informado, deve ser valido.

2) MODELO CERTIFICADOS.xlsx  →  aba Emissao em Massa > botao Procurar...
   A Nome* | B NR* | C Data do Treinamento (dd/mm/aaaa)*
   NRs disponiveis: {nrs}
   Funcionario inexistente e cadastrado automaticamente (sem CPF gera so o registro).

3) MODELO CARTOES BLOQUEIO.xlsx  →  aba Cartoes de Bloqueio > botao Importar Excel
   A Nome* | B CPF
   Marca na lista os funcionarios encontrados (por CPF ou nome exato).

4) MODELO ASO.xlsx  →  aba ASO > botao Importar Excel
   A Nome* | B CPF | C Tipo de ASO* | D Data do Exame (dd/mm/aaaa)* | E Validade em meses (1-120, vazio = 12)
   Tipos validos: {", ".join(ASO_TIPOS)}
   O funcionario precisa estar cadastrado. Os PDFs sao gerados automaticamente.

Dicas gerais:
- Linhas com erro NAO param a importacao — o app mostra o resumo com os detalhes.
- Datas podem ser digitadas como dd/mm/aaaa ou no formato de data do Excel.
- Campos vazios nas colunas opcionais sao aceitos normalmente.
"""
    (PASTA / "LEIA-ME.txt").write_text(leia_me, encoding="utf-8")

    for f in sorted(PASTA.iterdir()):
        print(f"[OK] {f.name}")


if __name__ == "__main__":
    main()
