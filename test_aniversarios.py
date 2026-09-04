"""Testes de data de nascimento + aniversariantes (v1.10.0, roadmap 2.15).

Cobre: validator do model, CRUD no repo (create/update/limpar),
get_aniversariantes (mes/dia), import Excel coluna E e export com a nova coluna.

Rodar: python test_aniversarios.py
"""

import sys
import tempfile
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

PASSOS = []


def check(nome, cond):
    PASSOS.append((nome, bool(cond)))
    print(f"[{'OK' if cond else 'FALHOU'}] {nome}")


# ── 1. Model: validator ──────────────────────────────────────

def test_model_validator():
    from src.core.models import Employee
    e = Employee(nome="Joao Pedro", data_nascimento="15/03/1990")
    check("dd/mm/aaaa normalizado p/ ISO", e.data_nascimento == "1990-03-15")
    e2 = Employee(nome="Maria Silva", data_nascimento="1990-03-15")
    check("ISO aceito direto", e2.data_nascimento == "1990-03-15")
    e3 = Employee(nome="Sem Data", data_nascimento="")
    check("vazio vira None", e3.data_nascimento is None)
    try:
        Employee(nome="Invalido", data_nascimento="31/02/1990")
        check("data impossivel rejeitada", False)
    except Exception:
        check("data impossivel rejeitada", True)


# ── 2. Repo: CRUD + aniversariantes ──────────────────────────

def test_repo(tmp: Path):
    from src.core.employee_repo import EmployeeRepository
    repo = EmployeeRepository(db_path=tmp / "aniv.db")

    id_joao = repo.create("Joao Pedro", None, data_nascimento="15/03/1990")
    id_maria = repo.create("Maria Silva", None, data_nascimento="15/09/1985")
    id_beto = repo.create("Beto Lima", None, data_nascimento="02/09/1978")
    id_sem = repo.create("Sem Nascimento", None)

    todos = {e.id: e for e in repo.get_all()}
    check("create grava ISO", todos[id_joao].data_nascimento == "1990-03-15")
    check("sem nascimento -> None", todos[id_sem].data_nascimento is None)

    hoje = date.today()
    # mes garantidamente DIFERENTE do atual (para nao poluir os filtros)
    outro_mes = (hoje.month % 12) + 1
    repo.update(id_joao, "Joao Pedro", data_nascimento=f"15/{outro_mes:02d}/1990")
    repo.update(id_maria, "Maria Silva", data_nascimento=f"20/{outro_mes:02d}/1985")
    # aniversariantes do mes corrente: dia 01 e dia de hoje
    mes_atual = f"{hoje.month:02d}"
    repo.update(id_beto, "Beto Lima", data_nascimento=f"01/{mes_atual}/1978")
    repo.update(id_sem, "Sem Nascimento", data_nascimento=f"{hoje.day:02d}/{mes_atual}/1999")

    do_mes = repo.get_aniversariantes(hoje.month)
    check("get_aniversariantes(mes) filtra o mes",
          {e.nome for e in do_mes} == {"Beto Lima", "Sem Nascimento"})
    check("ordenacao por dia",
          [e.nome for e in do_mes] == ["Beto Lima", "Sem Nascimento"])

    do_dia = repo.get_aniversariantes(hoje.month, hoje.day)
    check("aniversariante do dia exato",
          any(e.nome == "Sem Nascimento" for e in do_dia))

    # update sem flag NAO apaga (compat chamadas antigas)
    repo.update(id_joao, "Joao Pedro", telefone="21984209236")
    joao = [e for e in repo.get_all() if e.id == id_joao][0]
    check("valor preservado apos update sem campo",
          joao.data_nascimento == f"1990-{outro_mes:02d}-15")

    # limpar explicito
    repo.update(id_joao, "Joao Pedro", data_nascimento=None, limpar_nascimento=True)
    joao = [e for e in repo.get_all() if e.id == id_joao][0]
    check("limpar_nascimento apaga", joao.data_nascimento is None)

    # troca de data
    repo.update(id_maria, "Maria Silva", data_nascimento=f"20/{mes_atual}/1985")
    maria = [e for e in repo.get_all() if e.id == id_maria][0]
    check("update troca data", maria.data_nascimento == f"1985-{mes_atual}-20")


# ── 3. Import Excel coluna E ─────────────────────────────────

def test_importer(tmp: Path):
    import openpyxl
    from src.core.employee_repo import EmployeeRepository
    from src.utils.excel_importer import import_employees_from_excel, _parse_nascimento

    check("parser dd/mm/aaaa", _parse_nascimento("15/03/1990") == "1990-03-15")
    check("parser ISO", _parse_nascimento("1990-03-15") == "1990-03-15")
    check("parser datetime nativo",
          _parse_nascimento(datetime(1990, 3, 15)) == "1990-03-15")
    check("parser vazio -> None", _parse_nascimento(None) is None)
    try:
        _parse_nascimento("marco/1990")
        check("parser invalido gera erro", False)
    except ValueError:
        check("parser invalido gera erro", True)

    xlsx = tmp / "importa.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Nome", "CPF", "Funcao", "Telefone", "Nascimento"])
    ws.append(["Ana Souza", None, "Eletricista", None, "10/05/1992"])
    ws.append(["Bruno Alves", None, None, None, None])
    ws.append(["Carla Dias", None, None, None, "99/99/9999"])
    wb.save(xlsx)

    repo = EmployeeRepository(db_path=tmp / "imp.db")
    criados, duplicados, erros, detalhes = import_employees_from_excel(str(xlsx), repo)

    check("2 importados, 1 erro", criados == 2 and erros == 1)
    check("detalhe aponta data invalida",
          any("data de nascimento invalida" in d for d in detalhes))
    ana = [e for e in repo.get_all() if e.nome == "Ana Souza"][0]
    check("nascimento importado ISO", ana.data_nascimento == "1992-05-10")
    bruno = [e for e in repo.get_all() if e.nome == "Bruno Alves"][0]
    check("linha sem nascimento ok", bruno.data_nascimento is None)


# ── 4. Export com nova coluna ────────────────────────────────

def test_exporter(tmp: Path):
    import openpyxl
    from src.core.employee_repo import EmployeeRepository
    from src.utils.excel_exporter import export_employees_to_excel

    repo = EmployeeRepository(db_path=tmp / "exp.db")
    repo.create("Joao Pedro", None, data_nascimento="15/03/1990")

    out = tmp / "exporta.xlsx"
    n = export_employees_to_excel(repo, str(out))
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    check("coluna Data Nascimento no header", "Data Nascimento" in header)
    row2 = [c.value for c in ws[2]]
    idx = header.index("Data Nascimento")
    check("exportado como dd/mm/aaaa", row2[idx] == "15/03/1990")
    check("quantidade exportada", n == 1)


def main():
    with tempfile.TemporaryDirectory(prefix="normatech_aniv_", ignore_cleanup_errors=True) as td:
        tmp = Path(td)
        test_model_validator()
        test_repo(tmp)
        test_importer(tmp)
        test_exporter(tmp)

    falhas = [n for n, ok in PASSOS if not ok]
    print(f"\n{len(PASSOS) - len(falhas)}/{len(PASSOS)} testes OK")
    if falhas:
        print("FALHARAM:", falhas)
        sys.exit(1)


if __name__ == "__main__":
    main()
